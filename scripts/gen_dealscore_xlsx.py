"""
Gera _docs/Projeto/DealScore_Metodologia.xlsx a partir das regras
em dealscore/deal_score_rules.py + dealscore/client_priority.py.

Fonte única de verdade: as regras em código. Rode após qualquer
alteração nos pesos.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dealscore import deal_score_rules as R
from dealscore.client_priority import (
    WEIGHTS as CW,
    SEGMENTO_CLIENT_SCORES,
    DIAS_AGENDAR_BUCKETS,
    CARGOS_DECISOR,
    FIELDS as CFIELDS,
)

OUT = Path(__file__).resolve().parent.parent / "_docs" / "Projeto" / "DealScore_Metodologia.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CLIENT_FILL = PatternFill("solid", fgColor="FEF3C7")  # amarelo claro destaque cliente


def _autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _write_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")


def _write_dict_sheet(wb, name, data: dict, label="Item", tag=""):
    ws = wb.create_sheet(name)
    _write_header(ws, [label, "Pontos", "Tag"])
    row = 2
    for k, v in data.items():
        ws.cell(row=row, column=1, value=str(k))
        ws.cell(row=row, column=2, value=v)
        ws.cell(row=row, column=3, value=tag)
        row += 1
    _autosize(ws)


def _write_buckets_sheet(wb, name, buckets, unit, label):
    ws = wb.create_sheet(name)
    _write_header(ws, [f"≤ {unit}", "Pontos", "Descrição"])
    for i, (cap, pts) in enumerate(buckets, 2):
        ws.cell(row=i, column=1, value=cap)
        ws.cell(row=i, column=2, value=pts)
        ws.cell(row=i, column=3, value=label)
    _autosize(ws)


def main():
    wb = Workbook()
    # Capa / overview
    ws = wb.active
    ws.title = "Visão Geral"
    _write_header(ws, ["Categoria", "Faixa de Pontos", "Observações"])
    rows = [
        ("Stage (etapa)", "0 → +90", "Score fixo por estágio. Contato 1 usa SLA."),
        ("Funil (TOFU/MOFU/BOFU)", "+5 → +20", "Detectado por prefixo ix.T/M/B"),
        ("Atividades", "-10 → +10", "Engajamento real do deal"),
        ("Cargo (Person)", "-10 → +25", "CEO/Dono pesa mais"),
        ("Segmento (Person)", "0 → +15", "Tecnologia/Financeiro no topo"),
        ("Formato de vendas (Person)", "0 → +15", "Recorrência > venda única"),
        ("Plataforma (Deal)", "0 → +6", "Hotmart > Eduzz > Sympla"),
        ("Qualidade (site/email/phone/quest)", "-10 → +20", "Dados validados"),
        ("Estagnação", "0 → -120", "Penaliza inatividade"),
        ("Cliente (PRIORIDADE)", "0 → +120", "⭐ Pesos altos definidos pelo time comercial"),
        ("Override Lost", "= -100", "Status=lost força -100"),
        ("Override Quarentena", "= +5", "Stage Quarentena força +5"),
        ("Cap global", "[-100, 300]", "Limites finais"),
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            if "PRIORIDADE" in str(r[0]):
                c.fill = CLIENT_FILL
                c.font = Font(bold=True)
    _autosize(ws)

    # Cliente priority — destaque
    ws = wb.create_sheet("⭐ Cliente Priority")
    _write_header(ws, ["#", "Critério (cliente)", "Peso máx (orig.)", "Peso aplicado (2x)", "ENV override", "Field ID configurado"])
    cli_rows = [
        (1, "Lead Indicado",         "+30", CW["lead_indicado"],     "CLIENT_W_LEAD_INDICADO",     CFIELDS["lead_indicado"] or "fallback: Fonte=indicação"),
        (2, "CRM Preferencial",      "+20", CW["crm_preferencial"],  "CLIENT_W_CRM_PREFERENCIAL",  CFIELDS["crm_preferencial"] or "(definir)"),
        (3, "Qtd Pessoas no MKT",    "+20", CW["pessoas_mkt_max"],   "CLIENT_W_PESSOAS_MKT",       CFIELDS["pessoas_mkt"] or "(definir)"),
        (4, "Dias para Agendar",     "+20", CW["dias_agendar_max"],  "CLIENT_W_DIAS_AGENDAR",      "calc: add_time → stage 47"),
        (5, "Segmento",              "+10", CW["segmento_max"],      "CLIENT_W_SEGMENTO",          "person.segmento"),
        (6, "Qtd Pessoas Comercial", "+10", CW["pessoas_com_max"],   "CLIENT_W_PESSOAS_COMERCIAL", CFIELDS["pessoas_com"] or "(definir)"),
        (7, "Valores OK? (Não)",     "-20", CW["valores_nao_ok"],    "CLIENT_W_VALORES_NAO_OK",    CFIELDS["valores_ok"] or "(definir)"),
        (8, "Plataforma Integrada (Não)", "-10", CW["plataforma_nao_ok"], "CLIENT_W_PLATAFORMA_NAO_OK", CFIELDS["plataforma_integrada"] or "(definir)"),
        (9, "Cargo Decisor (Não)",   "-10", CW["cargo_nao_decisor"], "CLIENT_W_CARGO_NAO_DECISOR", "derivado: cargo ∉ decisores"),
    ]
    for i, r in enumerate(cli_rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = CLIENT_FILL
    note = ws.cell(row=len(cli_rows) + 4, column=1,
                   value="📋 Pesos derivados da planilha do cliente (anexo 23/04). Aplicado 2x sobre os pesos originais "
                         "para garantir prioridade. Critérios sem field_id configurado são no-op até definir CLIENT_FIELD_*.")
    note.font = Font(italic=True, color="92400E")
    _autosize(ws)

    # CRM Weights (calibrado por win-rate último ano)
    from dealscore.client_priority import CRM_WEIGHTS
    ws = wb.create_sheet("CRM Weights (Cliente)")
    _write_header(ws, ["CRM normalizado", "Peso aplicado", "Observação"])
    obs = {
        "pipedrive":"top win-rate (36%) + cliente prefere",
        "hubspot":"win-rate 35%",
        "moskit":"cliente prioriza (n pequeno)",
        "rdstation":"win-rate 23%",
        "agendor":"medium (n pequeno)",
        "salesforce":"medium",
        "kommo":"win-rate 21%",
        "zoho":"poucos dados",
        "bitrix":"poucos dados",
        "ploomes":"poucos dados",
        "fleeg":"poucos dados",
        "meetz":"poucos dados",
        "outro":"qualquer string ≠ vazio",
        "planilha":"usa planilha/excel",
        "whatsapp":"usa WhatsApp",
        "sem_crm":"declarou 'sem CRM' (win 4.5%)",
        "vazio":"campo vazio (default)",
    }
    for i,(k,v) in enumerate(sorted(CRM_WEIGHTS.items(), key=lambda x:-x[1]), 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=obs.get(k,""))
    _autosize(ws)

    # Segmento cliente (mapeamento próprio)
    ws = wb.create_sheet("Segmento (Cliente)")
    _write_header(ws, ["Segmento", "Pontos (escalados)"])
    for i, (k, v) in enumerate(sorted(SEGMENTO_CLIENT_SCORES.items(), key=lambda x: -x[1]), 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _autosize(ws)

    # Dias para agendar buckets
    ws = wb.create_sheet("Dias para Agendar (Cliente)")
    _write_header(ws, ["≤ dias", "% do peso máx", "Pontos"])
    for i, (cap, frac) in enumerate(DIAS_AGENDAR_BUCKETS, 2):
        ws.cell(row=i, column=1, value=cap)
        ws.cell(row=i, column=2, value=f"{int(frac*100)}%")
        ws.cell(row=i, column=3, value=int(round(CW["dias_agendar_max"] * frac)))
    _autosize(ws)

    # Cargos decisor
    ws = wb.create_sheet("Cargos Decisor (Cliente)")
    _write_header(ws, ["Cargo (presença = neutro)"])
    for i, c in enumerate(sorted(CARGOS_DECISOR), 2):
        ws.cell(row=i, column=1, value=c)
    ws.cell(row=len(CARGOS_DECISOR) + 3, column=1,
            value=f"Outros cargos preenchidos → penalidade {CW['cargo_nao_decisor']}").font = Font(italic=True)
    _autosize(ws)

    # Demais abas
    _write_dict_sheet(wb, "Stages", {f"stage_{k}": v for k, v in R.STAGE_SCORES.items()}, "Stage")
    _write_buckets_sheet(wb, "Stage Contato 1 SLA", R.CONTACT1_SLA_BUCKETS, "horas", "SLA primeiro contato")
    _write_dict_sheet(wb, "Funil", R.FUNIL_SCORES, "Faixa do funil")
    _write_dict_sheet(wb, "Cargo", R.CARGO_SCORES, "Cargo Person")
    _write_dict_sheet(wb, "Segmento", R.SEGMENTO_SCORES, "Segmento Person")
    _write_dict_sheet(wb, "Formato Vendas", R.FORMATO_VENDAS_SCORES, "Formato Person")
    _write_dict_sheet(wb, "Plataforma", R.PLATAFORMA_SCORES, "Plataforma Deal")
    _write_dict_sheet(wb, "Qualidade", {
        "Site válido (Sim)": R.SITE_VALIDO_SCORES["Sim"],
        "Site válido (Não)": R.SITE_VALIDO_SCORES["Não"],
        "Email válido (Sim)": R.EMAIL_VALIDO_SCORES["Sim"],
        "Email válido (Não)": R.EMAIL_VALIDO_SCORES["Não"],
        "Email empresarial (Sim)": R.EMAIL_EMPRESARIAL_SCORES["Sim"],
        "Phone válido (Sim)": R.PHONE_VALIDO_SCORES["Sim"],
        "Phone válido (Não)": R.PHONE_VALIDO_SCORES["Não"],
        "Questionário (Sim)": R.QUESTIONARIO_SCORES["Sim"],
        "Questionário (Não)": R.QUESTIONARIO_SCORES["Não"],
    }, "Validação")
    _write_dict_sheet(wb, "Status", R.STATUS_SCORES, "Status")
    _write_buckets_sheet(wb, "Estagnação", R.STAGNATION_BUCKETS_DAYS, "dias", "Dias sem movimento")
    _write_buckets_sheet(wb, "Atividades", R.ACTIVITIES_COUNT_BUCKETS, "qtd", "# atividades total")

    # Caps anti-ilusão
    ws = wb.create_sheet("Caps anti-ilusão")
    _write_header(ws, ["≥ dias parado", "Cap máximo do score"])
    for i, (d, cap) in enumerate(R.STAGNATION_CAPS, 2):
        ws.cell(row=i, column=1, value=d)
        ws.cell(row=i, column=2, value=cap)
    _autosize(ws)

    # Emoji tiers
    ws = wb.create_sheet("Emojis")
    _write_header(ws, ["Faixa de score", "Emoji", "Significado"])
    tiers = [
        ("≤ 0", "🪨", "Morto / frio (lost ou penalizado)"),
        ("1 – 50", "🌱", "Semente — sinais baixos"),
        ("51 – 100", "🌿", "Crescendo — tem potencial"),
        ("101 – 200", "🌳", "Alta prioridade — trabalhar com foco"),
        ("> 200", "🍀", "Estratégico — quase ganho"),
    ]
    for i, r in enumerate(tiers, 2):
        for j, v in enumerate(r, 1):
            ws.cell(row=i, column=j, value=v)
    _autosize(ws)

    wb.save(OUT)
    print(f"✅ Excel gerado: {OUT}")
    print(f"   Tamanho: {OUT.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
